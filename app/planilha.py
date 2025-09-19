import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import logging

def formatar_planilha_excel(df: pd.DataFrame, buffer: io.BytesIO):
    """Formata e escreve DataFrame para buffer Excel com formatação."""
    try:
        if df.empty:
            wb = Workbook()
            ws = wb.active
            ws.title = "Imóveis"
            ws.append(["MATRICULA", "UF", "CIDADE", "BAIRRO", "ENDERECO", "STATUS", 
                      "PRECO", "AVALIACAO", "DESCONTO", "AREA_PRIVATIVA", "AREA_DO_TERRENO", 
                      "TIPO", "MODALIDADE", "DATA_DISPUTA", "FGTS", "FINANCIAMENTO"])
            ws.append(["Nenhum dado encontrado"])
            wb.save(buffer)
            return

        if len(df) == 0:
            logging.warning("DataFrame está vazio para exportação")
            return

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Imóveis')
            
            worksheet = writer.sheets['Imóveis']

            header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            currency_format = 'R$ #,##0.00'
            cell_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions

            for i, col_name in enumerate(df.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                
                for j, cell in enumerate(worksheet[column_letter], 1):
                    if j > 1: 
                        cell.alignment = cell_alignment
                        
                        if col_name in ['PRECO', 'AVALIACAO'] and cell.value is not None:
                            try:
                                if isinstance(cell.value, (int, float)) and cell.value > 0:
                                    cell.number_format = currency_format
                            except (ValueError, TypeError):
                                pass
                    
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max(max_length + 4, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logging.info(f"Planilha Excel formatada com sucesso. Linhas: {len(df)}")
        
    except Exception as e:
        logging.error(f"Erro ao formatar planilha Excel: {e}", exc_info=True)
        try:
            df.to_excel(buffer, index=False, sheet_name='Imóveis', engine='openpyxl')
        except Exception as e2:
            logging.error(f"Erro ao criar planilha simples: {e2}")
            raise